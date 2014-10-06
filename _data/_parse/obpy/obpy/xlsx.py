import sys
import os
import json
import copy
import logging

import requests
import openpyxl

_logger = logging.getLogger(__name__)

class Error(Exception):
    pass

def _equal_values(iterator):
    try:
        iterator = iter(iterator)
        first = next(iterator)
        return all(first == rest for rest in iterator)
    except StopIteration:
        return True

def _xlsx_configs_for_worksheet(worksheet, config):
    max_search_depth = config['max_header_search_depth']

    configs = []
    # enumerate through rows, trying to find header row
    for i, row in enumerate(worksheet.iter_rows()):
        if i > max_search_depth:
            raise Error({'message': 'failed to find header row'})
       
        headers = copy.copy(config['headers'])
        for j, cell in enumerate(row):
            header_to_remove = None
            cellstr = cell.value
            if isinstance(cellstr, unicode):
                cellstr = cellstr.encode('ascii','ignore')
            for header in headers:
                if cellstr is None or isinstance(cellstr, float):
                    break
                for pattern in header['patterns']:
                    if pattern.match(cellstr):
                        header.update({'column':j})
                        configs.append(header)
                        header_to_remove = header
                        _logger.info(
                            'matched worksheet header "{0}" of python type "{1}" to header type "{2}"'.format(
                                cellstr,
                                type(cellstr),
                                header['type']))
                        break
                if not (header_to_remove is None):
                    break
            if not (header_to_remove is None):
                headers.remove(header_to_remove)
        if len(headers) == 0:
            return (configs, i)

def parse(xlsx_path, config):
    '''Parse an exel (*.xlsx) file into a python object.
    Input
        xlsx_path -> path to excel file
        config    -> parsing configuration
    Return
        python object representing excel file data

    The parsing generally assumes that the excel file is setup as a one or more
    worksheets, where each worksheet is a collection of rows with headers 
    located somewhere near the top of the worksheet. 
    The 'config' controls how the excel file is parsed and should have the 
    following format.

    config = {
        'workbooks': [
            {
                'type': <output field name>,
                'patterns': [
                    <regular expression to match>,
                    <regular expression to match>,
                    ...
                ],
                'max_header_search_depth': <maximum row # to search for a column header>,
                'headers': [
                    {
                        'type': <output field name>,
                        'patterns': [
                            <regular expression to match>,
                            <regular expression to match>,
                            ...
                        ],
                        'process': <function to process data element>
                    },
                    {
                        'type': <output field name>,
                        'patterns': [
                            <regular expression to match>,
                            <regular expression to match>,
                            ...
                        ],
                        'process': <function to process data element>
                    },
                    ...
                ]
            },
            ...
        ]
    }

    Example:

    import xlsx
    import re
    import os

    def _remove_file_ext(input_data):
        # remove extension from filename
        return os.path.splitext(input_data)[0]

    config = {
        'workbooks': [
            {
                'type': 'track_annotation',
                'patterns': [
                    re.compile(r'mood training.*', flags=re.IGNORECASE)
                ],
                'max_header_search_depth': 4,
                'headers': [
                    {
                        'type': 'filename',
                        'patterns': [
                            re.compile(r'filename', flags=re.IGNORECASE)
                        ],
                        'process': _remove_file_ext
                    }
                ]
            }
        ]
    }

    print xlsx.parse('./test.xlsx', config)
    '''

    data = {
    }

    xlwb = openpyxl.load_workbook(xlsx_path, use_iterators=True)
    names = xlwb.get_sheet_names()
    # iterate through excel worksheets
    for wsname in names:
        wsconfig = None
        # find matching worksheet in config
        for candidate_wsconfig in config['worksheets']:
            for wspattern in candidate_wsconfig['patterns']:
                if wspattern.match(wsname):
                    wsconfig = candidate_wsconfig
                    break
            if not wsconfig is None:
                break

        if wsconfig is None:
            # TODO: can we add functionality to ignore unprocessed worksheets in an excel doc?
            raise Error({
                'message': 'could not find match for worksheet in config',
                'worksheet': wsname
            })
            _logger.info(
                'matched worksheet title "{0}" to worksheet type "{1}"'.format(
                wsname,
                wsconfig['type']))

        
        # iterate through excel columns to get titles
        worksheet = []
        xlws = xlwb.get_sheet_by_name(wsname)

        # find header configurations and row # for worksheet
        column_configs, pos = _xlsx_configs_for_worksheet(xlws, wsconfig)

        #iterate through worksheet rows and process them
        start_row = pos + 1
        for row_count, row in enumerate(xlws.iter_rows()):
            # skip rows until we get one row past the header
            if row_count < start_row:
                continue

            # iterate through each column in row
            entry = {}
            for column_config in column_configs:
                entry[column_config['type']] = column_config['process'](row[column_config['column']].value)

            # don't add things that are all none
            if len(filter(lambda x: x[1], entry.items())) > 0:
                worksheet.append(entry)

        data[wsconfig['type']] = worksheet
    return data

